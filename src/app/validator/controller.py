import logging
from openpyxl import load_workbook

from src.app.validator.validator import Validator

logger = logging.getLogger(__name__)


class ValidatorController:
    def __init__(self, validator: Validator, path: str):
        self.validator = validator
        self.path = path

    def limpar_aba(self, aba="Inconsistência", first_row: int = 2):
        # Carrega o arquivo Excel com openpyxl
        workbook = load_workbook(self.path)

        if aba in workbook.sheetnames:
            sheet = workbook[aba]

            sheet.delete_rows(first_row, sheet.max_row)

            workbook.save(self.path)
            # logger.info(f"A aba '{aba}' foi limpa com sucesso.")
        else:
            logger.error(f"A aba '{aba}' não foi encontrada no arquivo.")
            exit(500)

    def validate_input(self):
        # clean Inconsistências tab before validate
        self.limpar_aba()

        # clean Solução tab before validate
        self.limpar_aba(aba="Solução")

        error_list = []

        # Checking age
        logger.info("[VALIDATOR] Checking Patient Age")
        error = self.validator.check_patient_age()
        if error:
            error_list.append("ERROR")

        # Checking profissional type
        logger.info("[VALIDATOR] Checking Professional Type")
        error = self.validator.check_professional_type()
        if error:
            error_list.append("AVISO")

        # Checking profissional type
        logger.info("[VALIDATOR] Checking Professional Availability")
        error = self.validator.check_professional_availability()
        if error:
            error_list.append("AVISO")

        # Checking profissional type
        logger.info("[VALIDATOR] Checking Professional Prefered Age Range")
        error = self.validator.check_professional_has_age_range()
        if error:
            error_list.append("AVISO")

        # Checking schedule profissional
        logger.info("[VALIDATOR] Checking Schedule Profissional")
        error = self.validator.check_has_schedule_profissional()
        if error:
            error_list.append("AVISO")

        # Checking schedule profissional
        logger.info("[VALIDATOR] Checking Local Profissional")
        error = self.validator.check_has_places_profissional()
        if error:
            error_list.append("ERROR")

        # Checking schedule profissional
        logger.info("[VALIDATOR] Checking Unique Profissional List")
        error = self.validator.check_same_professionals()
        if error:
            error_list.append("ERROR")

        # Checking schedule patient
        logger.info("[VALIDATOR] Checking Schedule Patient")
        error = self.validator.check_has_schedule_patient()
        if error:
            error_list.append("AVISO")

        # Checking schedule profissional
        logger.info("[VALIDATOR] Checking Local Patient")
        error = self.validator.check_has_places_patient()
        if error:
            error_list.append("ERROR")

        # Checking schedule profissional
        logger.info("[VALIDATOR] Checking Unique Patients List")
        error = self.validator.check_same_patients()
        if error:
            error_list.append("ERROR")

        return error_list
